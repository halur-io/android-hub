"""
Generic export utility for admin data export to CSV and Excel formats.
Supports flexible column mapping and formatting for different data models.
"""

import csv
import io
from datetime import date, datetime
from flask import Response, make_response


def export_to_csv(data, columns, filename='export.csv'):
    """
    Export data to CSV format.
    
    Args:
        data: List of dictionaries or SQLAlchemy model instances
        columns: List of dicts with 'field', 'header', and optional 'formatter'
            Example: [
                {'field': 'id', 'header': 'ID'},
                {'field': 'name', 'header': 'Name'},
                {'field': 'created_at', 'header': 'Created', 'formatter': lambda x: x.strftime('%Y-%m-% d') if x else ''}
            ]
        filename: Output filename
    
    Returns:
        Flask Response object with CSV content
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header row
    headers = [col['header'] for col in columns]
    writer.writerow(headers)
    
    # Write data rows
    for item in data:
        row = []
        for col in columns:
            # Get field value
            if isinstance(item, dict):
                value = item.get(col['field'])
            else:
                value = getattr(item, col['field'], None)
            
            # Apply formatter if provided
            if 'formatter' in col and col['formatter']:
                value = col['formatter'](value)
            else:
                # Default formatting
                value = format_value(value)
            
            row.append(value)
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'  # UTF-8 with BOM for Excel
    
    return response


def export_to_excel(data, columns, filename='export.xlsx'):
    """
    Export data to Excel format.
    
    Args:
        data: List of dictionaries or SQLAlchemy model instances
        columns: List of dicts with 'field', 'header', and optional 'formatter'
        filename: Output filename
    
    Returns:
        Flask Response object with Excel content
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(data, columns, filename.replace('.xlsx', '.csv'))
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'
    
    # Style for header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write header row
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col['header'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_idx, item in enumerate(data, start=2):
        for col_idx, col in enumerate(columns, start=1):
            # Get field value
            if isinstance(item, dict):
                value = item.get(col['field'])
            else:
                value = getattr(item, col['field'], None)
            
            # Apply formatter if provided
            if 'formatter' in col and col['formatter']:
                value = col['formatter'](value)
            else:
                # Default formatting
                value = format_value(value)
            
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(col['header']))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            cell_value = str(row[0].value) if row[0].value else ''
            max_length = max(max_length, len(cell_value))
        adjusted_width = min(max_length + 2, 50)  # Max width 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response


def format_value(value):
    """
    Default value formatter for common types.
    """
    if value is None:
        return ''
    elif isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
    elif isinstance(value, bool):
        return 'Yes' if value else 'No'
    elif isinstance(value, (int, float)):
        return str(value)
    else:
        return str(value)


def build_export_response(data, columns, format_type='csv', filename_base='export'):
    """
    Helper function to build export response based on format type.
    
    Args:
        data: Data to export
        columns: Column definitions
        format_type: 'csv' or 'excel'/'xlsx'
        filename_base: Base filename without extension
    
    Returns:
        Flask Response
    """
    if format_type.lower() in ['excel', 'xlsx']:
        return export_to_excel(data, columns, f'{filename_base}.xlsx')
    else:
        return export_to_csv(data, columns, f'{filename_base}.csv')


def export_to_excel_multi_sheet(sheets, filename='export.xlsx'):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return export_to_csv(sheets[0]['data'], sheets[0]['columns'], filename.replace('.xlsx', '.csv'))

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for idx, sheet_def in enumerate(sheets):
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = sheet_def.get('title', f'Sheet {idx + 1}')
        if sheet_def.get('rtl', False):
            ws.sheet_view.rightToLeft = True
        columns = sheet_def['columns']
        data = sheet_def['data']

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col['header'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, item in enumerate(data, start=2):
            for col_idx, col in enumerate(columns, start=1):
                if isinstance(item, dict):
                    value = item.get(col['field'])
                else:
                    value = getattr(item, col['field'], None)
                if 'formatter' in col and col['formatter']:
                    value = col['formatter'](value)
                else:
                    value = format_value(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, col in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(col['header']))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                cell_value = str(row[0].value) if row[0].value else ''
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
